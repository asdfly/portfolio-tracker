#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 专业报告生成器 — 多 Sheet 投资组合分析报告
"""
import sqlite3
import logging
from datetime import datetime
from typing import Dict, List, Optional
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class Styles:
    """Excel 样式常量"""
    TITLE_FONT = Font(name='微软雅黑', size=16, bold=True, color='FF1F4E79')
    SUBTITLE_FONT = Font(name='微软雅黑', size=11, color='FF666666')
    DATE_FONT = Font(name='微软雅黑', size=10, color='FF999999')
    HEADER_FONT = Font(name='微软雅黑', size=10, bold=True, color='FFFFFFFF')
    HEADER_FILL = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    DATA_FONT = Font(name='微软雅黑', size=10)
    DATA_ALIGN = Alignment(vertical='center')
    CENTER = Alignment(horizontal='center', vertical='center')
    RIGHT = Alignment(horizontal='right', vertical='center')
    UP_FONT = Font(name='微软雅黑', size=10, color='FF22C55E', bold=True)
    DOWN_FONT = Font(name='微软雅黑', size=10, color='FFEF4444', bold=True)
    NEUTRAL_FONT = Font(name='微软雅黑', size=10, color='FF666666')
    ALT_FILL = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')
    WHITE_FILL = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
    CARD_LABEL = Font(name='微软雅黑', size=9, color='FF666666')
    CARD_VALUE = Font(name='微软雅黑', size=14, bold=True, color='FF1F4E79')
    CARD_GREEN = Font(name='微软雅黑', size=14, bold=True, color='FF22C55E')
    CARD_RED = Font(name='微软雅黑', size=14, bold=True, color='FFEF4444')
    THIN_BORDER = Border(
        left=Side(style='thin', color='FFD9D9D9'),
        right=Side(style='thin', color='FFD9D9D9'),
        top=Side(style='thin', color='FFD9D9D9'),
        bottom=Side(style='thin', color='FFD9D9D9'),
    )


class ExcelReportGenerator:
    """Excel 多 Sheet 专业报告生成器"""

    def __init__(self, db_path: str):
        self.db_path = db_path

    def generate(self, output_path: str = None, date: str = None) -> str:
        """生成完整 Excel 报告，返回文件路径"""
        wb = Workbook()
        data = self._load_all_data(date)
        if data is None:
            raise ValueError("数据库中没有足够数据生成报告")

        self._write_summary_sheet(wb, data)
        self._write_positions_sheet(wb, data)
        self._write_returns_sheet(wb, data)
        self._write_risk_sheet(wb, data)
        self._write_technical_sheet(wb, data)
        self._write_alerts_sheet(wb, data)

        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        if output_path is None:
            report_date = data['report_date']
            output_path = str(
                Path(self.db_path).parent.parent.parent
                / 'data' / 'reports'
                / f"portfolio_report_{report_date}.xlsx"
            )
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info(f"Excel 报告已生成: {output_path}")
        return output_path

    def _load_all_data(self, date: str = None):
        """加载报告所需的全部数据"""
        conn = sqlite3.connect(self.db_path)
        try:
            if date is None:
                row = conn.execute("SELECT MAX(date) FROM portfolio_snapshots").fetchone()
                if not row or not row[0]:
                    return None
                date = row[0]

            positions = pd.read_sql_query(
                "SELECT * FROM portfolio_snapshots WHERE date = ? ORDER BY market_value DESC",
                conn, params=[date])
            if positions.empty:
                return None

            summary = pd.read_sql_query(
                "SELECT * FROM portfolio_summary ORDER BY date DESC", conn)

            codes = positions['code'].tolist()
            technical = pd.DataFrame()
            if codes:
                ph = ','.join(['?'] * len(codes))
                technical = pd.read_sql_query(
                    f"SELECT * FROM etf_technical WHERE code IN ({ph}) "
                    f"AND date = (SELECT MAX(date) FROM etf_technical) ORDER BY code",
                    conn, params=codes)

            alerts = pd.read_sql_query(
                "SELECT * FROM alerts ORDER BY created_at DESC LIMIT 50", conn)

            return {
                'report_date': date, 'positions': positions, 'summary': summary,
                'technical': technical, 'alerts': alerts,
                'latest_summary': summary.iloc[0].to_dict() if not summary.empty else {},
            }
        finally:
            conn.close()

    def _write_summary_sheet(self, wb, data):
        """Sheet 1: 组合概览"""
        ws = wb.active
        ws.title = '组合概览'
        ws.sheet_properties.tabColor = '4472C4'
        s = Styles
        latest = data['latest_summary']
        positions = data['positions']

        # 标题
        ws.merge_cells('A1:F1')
        ws['A1'] = '投资组合分析报告'
        ws['A1'].font = s.TITLE_FONT
        ws.row_dimensions[1].height = 40
        ws.merge_cells('A2:F2')
        ws['A2'] = f"报告日期: {data['report_date']}    生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = s.DATE_FONT
        ws.row_dimensions[2].height = 22

        # 核心指标
        tv = latest.get('total_value', 0)
        tc = latest.get('total_cost', 0)
        tp = latest.get('total_pnl', 0)
        tr = (tp / tc * 100) if tc > 0 else 0
        dp = latest.get('daily_pnl', 0)
        dr = latest.get('daily_return', 0)
        sharpe = latest.get('sharpe_ratio')
        max_dd = latest.get('max_drawdown')
        vol = latest.get('volatility')

        metrics = [
            ('总市值', f"\u00a5{tv:,.0f}", s.CARD_VALUE),
            ('累计盈亏', f"\u00a5{tp:+,.0f} ({tr:+.2f}%)", s.CARD_GREEN if tp >= 0 else s.CARD_RED),
            ('当日盈亏', f"\u00a5{dp:+,.0f} ({dr:+.2f}%)", s.CARD_GREEN if dp >= 0 else s.CARD_RED),
            ('持仓数量', f"{len(positions)} 只", s.CARD_VALUE),
            ('夏普比率', f"{sharpe:.2f}" if sharpe and sharpe == sharpe else "N/A",
             s.CARD_GREEN if sharpe and sharpe > 0 else s.CARD_RED if sharpe and sharpe < 0 else s.CARD_VALUE),
            ('最大回撤', f"{max_dd:.2f}%" if max_dd and max_dd == max_dd else "N/A",
             s.CARD_RED if max_dd and abs(max_dd) > 10 else s.CARD_GREEN),
            ('年化波动率', f"{vol:.2f}%" if vol and vol == vol else "N/A",
             s.CARD_RED if vol and vol > 25 else s.CARD_VALUE),
            ('盈/亏比', f"{latest.get('profit_count', 0)}/{latest.get('loss_count', 0)}", s.CARD_VALUE),
        ]

        row = 4
        ws.merge_cells('A4:F4')
        ws['A4'] = '核心指标'
        ws['A4'].font = Font(name='微软雅黑', size=12, bold=True, color='FF333333')

        for i, (label, value, font) in enumerate(metrics):
            col_off = (i % 2) * 3
            cr = row + 1 + (i // 2) * 2
            ws.cell(row=cr, column=1 + col_off, value=label).font = s.CARD_LABEL
            ws.cell(row=cr + 1, column=1 + col_off, value=value).font = font

        # 行业分布
        from config.settings import ETF_CATEGORIES, SECTOR_COLORS
        sector_data = {}
        for _, pos in positions.iterrows():
            code = str(pos['code'])
            sector = ETF_CATEGORIES.get(code, {}).get('sector', '其他')
            mv = pos.get('market_value', 0)
            pnl = pos.get('pnl', 0)
            if sector not in sector_data:
                sector_data[sector] = {'mv': 0, 'pnl': 0, 'count': 0}
            sector_data[sector]['mv'] += mv
            sector_data[sector]['pnl'] += pnl
            sector_data[sector]['count'] += 1

        sr = row + 1 + (len(metrics) // 2) * 2 + 2
        ws.merge_cells(f'A{sr}:F{sr}')
        ws[f'A{sr}'] = '行业分布'
        ws[f'A{sr}'].font = Font(name='微软雅黑', size=12, bold=True, color='FF333333')
        sr += 1

        for ci, h in enumerate(['行业', '市值', '占比', '盈亏', '收益率', '持仓数']):
            c = ws.cell(row=sr, column=ci + 1, value=h)
            c.font = s.HEADER_FONT; c.fill = s.HEADER_FILL; c.alignment = s.HEADER_ALIGN; c.border = s.THIN_BORDER
        sr += 1

        for si, (sn, sd) in enumerate(sorted(sector_data.items(), key=lambda x: x[1]['mv'], reverse=True)):
            mv = sd['mv']
            wt = mv / tv * 100 if tv > 0 else 0
            pnl = sd['pnl']
            pnl_r = (pnl / (mv - pnl) * 100) if (mv - pnl) > 0 else 0
            vals = [sn, mv, f"{wt:.1f}%", pnl, pnl_r, sd['count']]
            fill = s.ALT_FILL if si % 2 == 0 else s.WHITE_FILL
            for ci, val in enumerate(vals):
                c = ws.cell(row=sr, column=ci + 1, value=val)
                c.font = s.DATA_FONT; c.fill = fill; c.border = s.THIN_BORDER
                if ci == 0:
                    c.font = Font(name='微软雅黑', size=10, bold=True, color='FF' + SECTOR_COLORS.get(sn, '333333').lstrip('#'))
                elif ci in (1, 3):
                    c.number_format = '#,##0.00'; c.alignment = s.RIGHT
                    if ci == 3:
                        c.font = s.UP_FONT if val >= 0 else s.DOWN_FONT
                elif ci == 4:
                    c.number_format = '+0.00"%"'; c.alignment = s.RIGHT
                    c.font = s.UP_FONT if val >= 0 else s.DOWN_FONT
                else:
                    c.alignment = s.CENTER
            sr += 1

        for i, w in enumerate([16, 16, 10, 16, 12, 10]):
            ws.column_dimensions[get_column_letter(i + 1)].width = w

    def _write_positions_sheet(self, wb, data):
        """Sheet 2: 持仓明细"""
        ws = wb.create_sheet('持仓明细')
        ws.sheet_properties.tabColor = '22C55E'
        s = Styles
        positions = data['positions']
        tv = data['latest_summary'].get('total_value', 1)

        ws.merge_cells('A1:K1')
        ws['A1'] = f"持仓明细 \u2014 {data['report_date']}"
        ws['A1'].font = s.TITLE_FONT; ws.row_dimensions[1].height = 36

        headers = ['代码', '名称', '行业', '持仓量', '成本价', '现价',
                   '市值', '盈亏', '收益率', '占比', 'Beta']
        for ci, h in enumerate(headers):
            c = ws.cell(row=3, column=ci + 1, value=h)
            c.font = s.HEADER_FONT; c.fill = s.HEADER_FILL; c.alignment = s.HEADER_ALIGN; c.border = s.THIN_BORDER
        ws.row_dimensions[3].height = 28

        from config.settings import ETF_CATEGORIES
        for pi, (_, pos) in enumerate(positions.iterrows()):
            r = pi + 4
            code = str(pos['code'])
            sector = ETF_CATEGORIES.get(code, {}).get('sector', '其他')
            mv = pos.get('market_value', 0)
            pnl = pos.get('pnl', 0)
            pnl_rate = pos.get('pnl_rate', 0)
            weight = mv / tv * 100 if tv > 0 else 0
            vals = [code, pos.get('name', ''), sector, pos.get('quantity', 0),
                    pos.get('cost_price', 0), pos.get('current_price', 0),
                    mv, pnl, pnl_rate, weight, pos.get('beta', '')]
            fill = s.ALT_FILL if pi % 2 == 0 else s.WHITE_FILL
            for ci, val in enumerate(vals):
                c = ws.cell(row=r, column=ci + 1, value=val)
                c.font = s.DATA_FONT; c.fill = fill; c.border = s.THIN_BORDER
                if ci in (3, 4, 5, 6):
                    c.number_format = '#,##0.000'; c.alignment = s.RIGHT
                elif ci == 7:
                    c.number_format = '#,##0.00'; c.alignment = s.RIGHT
                    c.font = s.UP_FONT if val >= 0 else s.DOWN_FONT
                elif ci == 8:
                    c.number_format = '+0.00"%"'; c.alignment = s.RIGHT
                    c.font = s.UP_FONT if val >= 0 else s.DOWN_FONT
                elif ci == 9:
                    c.number_format = '0.0"%"'; c.alignment = s.RIGHT
                elif ci == 10:
                    c.number_format = '0.000'; c.alignment = s.CENTER; c.font = s.NEUTRAL_FONT
                else:
                    c.alignment = Alignment(horizontal='left', vertical='center')
        ws.freeze_panes = 'A4'
        ws.auto_filter.ref = f"A3:K{len(positions) + 3}"
        for i, w in enumerate([10, 20, 8, 10, 10, 10, 14, 14, 10, 8, 8]):
            ws.column_dimensions[get_column_letter(i + 1)].width = w

    def _write_returns_sheet(self, wb, data):
        """Sheet 3: 收益汇总"""
        ws = wb.create_sheet('收益汇总')
        ws.sheet_properties.tabColor = 'F59E0B'
        s = Styles
        summary = data['summary']

        ws.merge_cells('A1:H1')
        ws['A1'] = '收益汇总'; ws['A1'].font = s.TITLE_FONT; ws.row_dimensions[1].height = 36

        headers = ['日期', '总市值', '累计盈亏', '累计收益率', '当日盈亏', '日收益率', 'vs沪深300', '夏普比率']
        for ci, h in enumerate(headers):
            c = ws.cell(row=3, column=ci + 1, value=h)
            c.font = s.HEADER_FONT; c.fill = s.HEADER_FILL; c.alignment = s.HEADER_ALIGN; c.border = s.THIN_BORDER
        ws.row_dimensions[3].height = 28

        display_df = summary.head(60)
        for pi, (_, row) in enumerate(display_df.iterrows()):
            r = pi + 4
            tc = row.get('total_cost', 1)
            cum_ret = (row.get('total_pnl', 0) / tc * 100) if tc > 0 else 0
            vals = [row.get('date', ''), row.get('total_value', 0), row.get('total_pnl', 0),
                    cum_ret, row.get('daily_pnl', 0), row.get('daily_return', 0),
                    row.get('vs_hs300', 0), row.get('sharpe_ratio', '')]
            fill = s.ALT_FILL if pi % 2 == 0 else s.WHITE_FILL
            for ci, val in enumerate(vals):
                c = ws.cell(row=r, column=ci + 1, value=val)
                c.font = s.DATA_FONT; c.fill = fill; c.border = s.THIN_BORDER
                if ci == 0:
                    c.number_format = 'YYYY-MM-DD'; c.alignment = s.CENTER
                elif ci in (1, 2, 4):
                    c.number_format = '#,##0.00'; c.alignment = s.RIGHT
                    if ci in (2, 4):
                        c.font = s.UP_FONT if val >= 0 else s.DOWN_FONT
                elif ci in (3, 5, 6):
                    c.number_format = '+0.00"%"'; c.alignment = s.RIGHT
                    c.font = s.UP_FONT if val >= 0 else s.DOWN_FONT
                elif ci == 7:
                    c.number_format = '0.00'; c.alignment = s.CENTER
        ws.freeze_panes = 'A4'
        ws.auto_filter.ref = f"A3:H{len(display_df) + 3}"
        for i, w in enumerate([12, 14, 14, 12, 14, 12, 12, 10]):
            ws.column_dimensions[get_column_letter(i + 1)].width = w

    def _write_risk_sheet(self, wb, data):
        """Sheet 4: 风险报告"""
        ws = wb.create_sheet('风险报告')
        ws.sheet_properties.tabColor = 'EF4444'
        s = Styles
        latest = data['latest_summary']
        summary = data['summary']
        positions = data['positions']

        ws.merge_cells('A1:F1')
        ws['A1'] = '风险分析报告'; ws['A1'].font = s.TITLE_FONT; ws.row_dimensions[1].height = 36

        # 风险概览
        ws.merge_cells('A3:F3')
        ws['A3'] = '风险概览'
        ws['A3'].font = Font(name='微软雅黑', size=12, bold=True, color='FF333333')
        risk_items = [
            ('夏普比率', latest.get('sharpe_ratio'), lambda v: '优秀' if v and v > 1 else '良好' if v and v > 0.5 else '一般'),
            ('最大回撤', latest.get('max_drawdown'), lambda v: '需警惕' if v and abs(v) > 15 else '可控'),
            ('年化波动率', latest.get('volatility'), lambda v: '偏高' if v and v > 25 else '适中'),
        ]
        r = 4
        for label, val, ev in risk_items:
            ws.cell(row=r, column=1, value=label).font = s.CARD_LABEL
            vs = f"{val:.2f}" if val and val == val else "N/A"
            ws.cell(row=r, column=2, value=vs).font = s.CARD_VALUE
            es = ev(val)
            ec = ws.cell(row=r, column=3, value=es)
            ec.font = Font(name='微软雅黑', size=10,
                          color='FFEF4444' if '警惕' in es or '偏高' in es else 'FF22C55E' if '优' in es or '良好' in es or '可控' in es or '适中' in es else 'FF666666')
            r += 1

        # Beta 分布
        r += 1
        ws.merge_cells(f'A{r}:F{r}')
        ws[f'A{r}'] = '个股 Beta 分布'
        ws[f'A{r}'].font = Font(name='微软雅黑', size=12, bold=True, color='FF333333')
        r += 1
        for ci, h in enumerate(['代码', '名称', '行业', 'Beta', '风险等级']):
            c = ws.cell(row=r, column=ci + 1, value=h)
            c.font = s.HEADER_FONT; c.fill = s.HEADER_FILL; c.alignment = s.HEADER_ALIGN; c.border = s.THIN_BORDER
        r += 1
        from config.settings import ETF_CATEGORIES
        for pi, (_, pos) in enumerate(positions.iterrows()):
            code = str(pos['code'])
            beta = pos.get('beta', 0)
            rl = '高' if beta and beta == beta and abs(beta) > 1.3 else '中' if beta and beta == beta else '低'
            fill = s.ALT_FILL if pi % 2 == 0 else s.WHITE_FILL
            for ci, val in enumerate([code, pos.get('name', ''), ETF_CATEGORIES.get(code, {}).get('sector', ''), beta, rl]):
                c = ws.cell(row=r, column=ci + 1, value=val)
                c.font = s.DATA_FONT; c.fill = fill; c.border = s.THIN_BORDER
                if ci == 3:
                    c.number_format = '0.000'; c.alignment = s.CENTER
                elif ci == 4:
                    c.alignment = s.CENTER
                    if val == '高': c.font = Font(name='微软雅黑', size=10, color='FFEF4444', bold=True)
                    elif val == '低': c.font = Font(name='微软雅黑', size=10, color='FF22C55E')
            r += 1

        # 近期趋势
        r += 1
        ws.merge_cells(f'A{r}:D{r}')
        ws[f'A{r}'] = '近期风险指标趋势'
        ws[f'A{r}'].font = Font(name='微软雅黑', size=12, bold=True, color='FF333333')
        r += 1
        for ci, h in enumerate(['日期', '夏普比率', '最大回撤(%)', '年化波动率(%)']):
            c = ws.cell(row=r, column=ci + 1, value=h)
            c.font = s.HEADER_FONT; c.fill = s.HEADER_FILL; c.alignment = s.HEADER_ALIGN; c.border = s.THIN_BORDER
        r += 1
        for pi, (_, row) in enumerate(summary.head(30).iterrows()):
            fill = s.ALT_FILL if pi % 2 == 0 else s.WHITE_FILL
            for ci, val in enumerate([row.get('date', ''), row.get('sharpe_ratio', ''), row.get('max_drawdown', ''), row.get('volatility', '')]):
                c = ws.cell(row=r, column=ci + 1, value=val if val == val else '')
                c.font = s.DATA_FONT; c.fill = fill; c.border = s.THIN_BORDER
                if ci == 0: c.number_format = 'YYYY-MM-DD'; c.alignment = s.CENTER
                else: c.number_format = '0.00'; c.alignment = s.CENTER
            r += 1
        for i, w in enumerate([12, 20, 8, 10, 10, 10]):
            ws.column_dimensions[get_column_letter(i + 1)].width = w

    def _write_technical_sheet(self, wb, data):
        """Sheet 5: 技术指标"""
        ws = wb.create_sheet('技术指标')
        ws.sheet_properties.tabColor = 'A855F7'
        s = Styles
        tech = data['technical']

        ws.merge_cells('A1:J1')
        ws['A1'] = f"技术指标概览 \u2014 {data['report_date']}"
        ws['A1'].font = s.TITLE_FONT; ws.row_dimensions[1].height = 36

        headers = ['代码', '名称', '行业', '趋势', '均线信号', 'MACD信号', 'RSI值', 'RSI状态', 'KDJ信号', '布林位置']
        for ci, h in enumerate(headers):
            c = ws.cell(row=3, column=ci + 1, value=h)
            c.font = s.HEADER_FONT; c.fill = s.HEADER_FILL; c.alignment = s.HEADER_ALIGN; c.border = s.THIN_BORDER
        ws.row_dimensions[3].height = 28

        from config.settings import ETF_CATEGORIES
        sig_c = {'多头排列': '22C55E', '金叉': '22C55E', '上涨': '22C55E',
                 '空头排列': 'EF4444', '死叉': 'EF4444', '下跌': 'EF4444'}
        for pi, (_, row) in enumerate(tech.iterrows()):
            r = pi + 4
            code = str(row.get('code', ''))
            cat = ETF_CATEGORIES.get(code, {})
            vals = [code, cat.get('name', code), cat.get('sector', '其他'),
                    row.get('trend', ''), row.get('ma_signal', ''), row.get('macd_signal', ''),
                    row.get('rsi_value', ''), row.get('rsi_status', ''),
                    row.get('kdj_signal', ''), row.get('bollinger_position', '')]
            fill = s.ALT_FILL if pi % 2 == 0 else s.WHITE_FILL
            for ci, val in enumerate(vals):
                v = val if val and val == val else ''
                c = ws.cell(row=r, column=ci + 1, value=v)
                c.font = s.DATA_FONT; c.fill = fill; c.border = s.THIN_BORDER; c.alignment = s.CENTER
                if ci >= 3 and isinstance(v, str):
                    for kw, color in sig_c.items():
                        if kw in v:
                            c.font = Font(name='微软雅黑', size=10, color=color, bold=True); break
                if ci == 0: c.alignment = Alignment(horizontal='left', vertical='center')
                if ci == 6: c.number_format = '0.0'
        ws.freeze_panes = 'A4'
        if len(tech) > 0: ws.auto_filter.ref = f"A3:J{len(tech) + 3}"
        for i, w in enumerate([10, 20, 8, 8, 12, 12, 8, 10, 12, 14]):
            ws.column_dimensions[get_column_letter(i + 1)].width = w

    def _write_alerts_sheet(self, wb, data):
        """Sheet 6: 告警记录"""
        ws = wb.create_sheet('告警记录')
        ws.sheet_properties.tabColor = 'F97316'
        s = Styles
        alerts = data['alerts']

        ws.merge_cells('A1:E1')
        ws['A1'] = '近期告警记录'; ws['A1'].font = s.TITLE_FONT; ws.row_dimensions[1].height = 36
        if alerts.empty:
            ws['A3'] = '暂无告警记录'; ws['A3'].font = s.NEUTRAL_FONT; return

        for ci, h in enumerate(['时间', '规则', '级别', '告警内容', '状态']):
            c = ws.cell(row=3, column=ci + 1, value=h)
            c.font = s.HEADER_FONT; c.fill = s.HEADER_FILL; c.alignment = s.HEADER_ALIGN; c.border = s.THIN_BORDER
        ws.row_dimensions[3].height = 28

        lv_c = {'error': 'EF4444', 'warning': 'F59E0B', 'info': '3B82F6'}
        for pi, (_, row) in enumerate(alerts.iterrows()):
            r = pi + 4
            level = row.get('level', 'info')
            fill = s.ALT_FILL if pi % 2 == 0 else s.WHITE_FILL
            vals = [row.get('created_at', ''), row.get('rule_name', ''), level,
                    row.get('message', ''), '已处理' if row.get('acknowledged') else '待处理']
            for ci, val in enumerate(vals):
                c = ws.cell(row=r, column=ci + 1, value=val)
                c.font = s.DATA_FONT; c.fill = fill; c.border = s.THIN_BORDER
                if ci == 0: c.number_format = 'YYYY-MM-DD HH:MM:SS'; c.alignment = s.CENTER
                elif ci == 2:
                    c.alignment = s.CENTER
                    c.font = Font(name='微软雅黑', size=10, color='FF' + lv_c.get(level, 'FF666666'), bold=True)
                elif ci == 4:
                    c.alignment = s.CENTER
                    c.font = Font(name='微软雅黑', size=10, color='FF22C55E' if val == '已处理' else 'FFEF4444', bold=(val == '待处理'))
                else: c.alignment = Alignment(horizontal='left', vertical='center')
        for i, w in enumerate([22, 18, 10, 40, 10]):
            ws.column_dimensions[get_column_letter(i + 1)].width = w


if __name__ == '__main__':
    import sys
    sys.path.insert(0, str(Path(__file__).parent.parent.parent))
    from config.settings import DATABASE_PATH
    gen = ExcelReportGenerator(str(DATABASE_PATH))
    out = gen.generate()
    print(f"报告已生成: {out}")
